"""
Script Name: Security_URL_Check.py
Author: Maurice Hofmann
Date: 2024-02-08
Version: 1.0.0

Description:
    This script validates and analyzes the security of URLs listed in CSV, TXT, or Excel files.
    It performs automated SSL/TLS checks using the "testssl.sh" tool, generates a detailed
    Excel report of the results, and optionally sends the report via email.

Dependencies:
    - pandas: For reading input files and generating structured reports.
    - numpy: For splitting datasets for parallel processing.
    - openpyxl: For writing Excel reports.
    - validators: For validating URLs and email addresses.
    - subprocess: For executing external commands (testssl.sh).
    - smtplib, email.message: For sending reports via email.
    - threading, multiprocessing: For parallel URL processing.
    - shutil, os, sys, json, time, platform, re: For file management and system operations.

Notes:
    - The script expects a file with URLs in the "Files_Security_Check" directory.
    - Configuration options such as "checkAll" and "receiver_mail" must be provided in "config.JSON".
    - On completion, the results are saved in an Excel report in the "Logfiles" directory.
    - The processed input file is automatically archived.
"""

#===============================
# Imports
#===============================

import time 
import ctypes
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



#===============================
# Variables
#===============================

optimization_number = 1
PGPT_URL = ""
FILE_PATH = "Example-Fragen-Automatisierung.xlsx"



#===============================
# Functions
#===============================

def open_pgpt(pgpt_url):
    """
    Opens a private GPT website in a Chrome browser window.

    This function initializes a Chrome WebDriver, navigates to the specified URL, 
    maximizes the browser window, and returns the WebDriver instance.

    Parameters:
    pgpt_url (str): The URL of the private GPT website to be opened.

    Returns:
    selenium.webdriver.Chrome: The WebDriver instance for the Chrome browser.

    Example:
    driver = open_pgpt("http:example.com/")
    """
     
    driver = webdriver.Chrome()
    driver.get(pgpt_url)

    return driver



def interact_pgpt(question, driver):
    """
    Interacts with a private GPT website by sending a question and retrieving the response.

    This function waits until a specific textarea element, identified by its XPath, is present on the page. 
    Once the element is located, it sends the provided question along with an ENTER key to simulate form submission. 
    After submitting the question, the function calls `get_pgpt_answer` to retrieve the answer from the website.

    Parameters:
    question (str): The question to be sent to the private GPT website.
    driver (selenium.webdriver.Chrome): The WebDriver instance used to interact with the private GPT website.

    Returns:
    str: The response from the private GPT website to the submitted question.

    Example:
    response = interact_pgpt("What is the capital of France?", driver)
    """

    ENTER = "\ue007"

    WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.XPATH, "/html/body/gradio-app/div/div/div[1]/div/div/div[2]/div[2]/div[3]/div/div[1]/div/label/textarea")))
    element = driver.find_element(By.XPATH, "/html/body/gradio-app/div/div/div[1]/div/div/div[2]/div[2]/div[3]/div/div[1]/div/label/textarea")
    
    element.send_keys(question + ENTER)

    answer = get_pgpt_answer(driver)

    return answer



def get_pgpt_answer(driver):
    """
    Retrieves the answer from a private GPT website using the provided WebDriver instance.

    This function waits until a specific element, identified by its XPath, is present on the page. 
    Once the element is located, it checks if the element's text contains the word "Sources". 
    If the word is found, the function waits for an additional 8 seconds before retrieving the text of the element 
    as the answer. The retrieved answer is then returned.

    Parameters:
    driver (selenium.webdriver.Chrome): The WebDriver instance used to interact with the private GPT website.

    Returns:
    str: The text content of the located element, which represents the answer from the private GPT website.

    Example:
    answer = get_pgpt_answer(driver)
    """

    WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.XPATH, "/html/body/gradio-app/div/div/div[1]/div/div/div[2]/div[2]/div[1]/div[2]/div/div/div[2]/div[2]/button")))

    element = driver.find_element(By.XPATH, "/html/body/gradio-app/div/div/div[1]/div/div/div[2]/div[2]/div[1]/div[2]/div/div/div[2]/div[2]/button")

    if(element.text.find("Sources")):
        time.sleep(8)
        answer = element.text
        
    return answer

    

def open_file(filepath):
    """
    Opens an Excel file and returns the workbook and active sheet.

    This function uses the openpyxl library to load an Excel workbook from the specified file path. 
    It then retrieves the active sheet from the workbook and returns both the workbook and the active sheet.

    Parameters:
    filepath (str): The path to the Excel file to be opened.

    Returns:
    tuple: A tuple containing the loaded workbook (openpyxl.workbook.workbook.Workbook) and the active sheet (openpyxl.worksheet.worksheet.Worksheet).

    Example:
    workbook, sheet = open_file("path/to/excel_file.xlsx")
    """

    excel = openpyxl.load_workbook(filepath)
    sheet = excel.active

    return excel, sheet



#===============================
# Main
#===============================

if __name__ == "__main__":
     
    optimization_number += 1

    pgpt_driver = open_pgpt(PGPT_URL)
    excel, excel_filesheet = open_file(FILE_PATH)
    

    for line in range(2, excel_filesheet.max_row + 1):

        # get quesition from column one
        question= excel_filesheet.cell(row = line, column = 1)
        print(question.value)

        answer = interact_pgpt(question.value, driver=pgpt_driver)

        # write answer of LLM in Excel
        excel_filesheet.cell(row=line, column=optimization_number).value = answer

        pgpt_driver.refresh()

    
    excel.save(FILE_PATH)
    
    ctypes.windll.user32.MessageBoxW(0, "Process Finished", "Successful", 1)